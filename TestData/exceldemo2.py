import openpyxl

book = openpyxl.load_workbook("D:\\Python stuff\\excel\\exceldemo.xlsx")


sheet = book.active
Dict = {}
for i in range(1,sheet.max_row+1):#to get rows
    if sheet.cell(row=i, column=1).value == "Testcase4":
        for j in range(1,sheet.max_column+1):#to get columns
            Dict[sheet.cell(row=1,column= j).value] = sheet.cell(row=i,column= j).value

print(Dict)