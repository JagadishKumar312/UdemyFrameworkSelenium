import openpyxl


class HomePageData:
    test_HomePageData = [{"firstname": "Jagdish", "email": "jaggu.career@gmail.com", "gender": "Male"},
                         {"firstname": "Kumar", "email": "jagadish.kumar@gmail.com", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("D:\\Python stuff\\excel\\exceldemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
